from tkinter import *
from openpyxl import *
root=Tk()
root.title("ClassRoom Pinboard")
root.geometry("1199x600+100+50")
root.resizable(False,False)
Frame1=Frame(root,bg="white")
Frame1.place(x=150,y=150,height=340,width=500)
def describe():
	extra=Label(Frame1,text="Select the day",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=250)
global extra	
lbl_user=Label(Frame1,text="Subject to be deleted",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=140)
txt_user=Entry(Frame1,font=("times new roman",15),bg="lightgray")
txt_user.place(x=90,y=170,width=350,height=35)
remove_button=Button(Frame1,text="Remove Subject",command=describe,cursor="hand2",bg="white",fg="#d77337",bd=0,font=("times new roman",12)).place(x=90,y=225)
my_menu=Menu(root)
root.config(menu=my_menu)
def mondayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/monday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			sheet.delete_rows(rowno)
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=250)
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/monday.xlsx')
def tuesdayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/tuesday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			sheet.delete_rows(rowno)
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=250)
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/tuesday.xlsx')	
def wednesdayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/wednesday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			sheet.delete_rows(rowno)
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=250)
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/wednesday.xlsx')	
def thursdayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/thursday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			sheet.delete_rows(rowno)
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=250)
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/thursday.xlsx')	
def fridayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/friday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			sheet.delete_rows(rowno)
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=250)
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/friday.xlsx')
def saturdayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/saturday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			sheet.delete_rows(rowno)
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=250)
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/saturday.xlsx')			
			
file_menu=Menu(my_menu)
my_menu.add_cascade(label="Days",menu=file_menu)
file_menu.add_command(label="Monday",command=mondayfun)
file_menu.add_command(label="Tuesday",command=tuesdayfun)
file_menu.add_command(label="Wednesday",command=wednesdayfun)
file_menu.add_command(label="Thursday",command=thursdayfun)
file_menu.add_command(label="Friday",command=fridayfun)
file_menu.add_command(label="Saturday",command=saturdayfun)


root.mainloop()