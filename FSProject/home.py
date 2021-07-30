from tkinter import *
from PIL import ImageTk 
import subprocess
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import *
root=Tk()
root.title("HomePage")
root.geometry("1199x600+100+50")
root.resizable(False,False)
bg=ImageTk.PhotoImage(file="sample1.jpg")
def add_link():
    cmd="python3 addlink.py"
    subprocess.Popen(cmd,shell=True)
def add_link1():
    cmd="python3 showlink.py"
    subprocess.Popen(cmd,shell=True) 
def add_link2():
    cmd="python3 update.py"
    subprocess.Popen(cmd,shell=True)
def add_link3():
    cmd="python3 delete.py"
    subprocess.Popen(cmd,shell=True)  
def searchfun():
	flag=0
	name=txt_user.get()
	if(txt_user1.get()=="monday"):
		wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/monday.xlsx')
	if(txt_user1.get()=="tuesday"):
		wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/tuesday.xlsx')
	if(txt_user1.get()=="wednesday"):
		wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/wednesday.xlsx')
	if(txt_user1.get()=="thursday"):
		wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/thursday.xlsx')
	if(txt_user1.get()=="friday"):
		wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/friday.xlsx')
	if(txt_user1.get()=="saturday"):
		wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/saturday.xlsx')
	else:
		extra=Label(Frame1,text="Invalid day",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=0,y=235)
	sheet = wb.active
	column_a=sheet['A']
	column_e=sheet['E']
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			flag=1

	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=0,y=235)
			
	else:
		for cell in column_e:
			if(cell.row==rowno):extra=Label(Frame1,text=cell.value,font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=0,y=235)
	if(txt_user1.get()=="monday"):
		
		wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/monday.xlsx')
	if(txt_user1.get()=="tuesday"):
		
		wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/tuesday.xlsx')	
	if(txt_user1.get()=="wednesday"):
		
		wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/wednesday.xlsx')
	if(txt_user1.get()=="thursday"):
		
		wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/thursday.xlsx')
	if(txt_user1.get()=="friday"):
		
		wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/friday.xlsx')
	if(txt_user1.get()=="saturday"):
		
		wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/saturday.xlsx')				
bg_image=Label(root,image=bg).place(x=0,y=0,relwidth=1,relheight=1)
Frame_name=Frame(root,bg="white")
Frame_name.place(x=0,y=0,height=40,width=1199)
Frame1=Frame(root,bg="white")
Frame1.place(x=150,y=160,height=340,width=500)

project_name=Label(Frame_name,text="ClassRoom Pinboard",font=("Lucida Calligraphy",20,"bold"),fg="black").place(x=500,y=5)
project_name1=Label(Frame_name,text="Welcome,1JS18IS001",font=("Lucida Calligraphy",20),fg="black").place(x=0,y=5)
OptionFrame=Frame(root)
OptionFrame.place(x=150,y=130,height=30,width=500)
add_button=Button(OptionFrame,text="Add Event",command=add_link,cursor="hand2",bg="white",fg="#d77337",bd=0,font=("times new roman",12)).place(x=0,y=0)
show_buttton=Button(OptionFrame,text="Show Event",command=add_link1,cursor="hand2",bg="white",fg="#d77337",bd=0,font=("times new roman",12)).place(x=70,y=0)
update_buttton=Button(OptionFrame,text="Update Event",command=add_link2,cursor="hand2",bg="white",fg="#d77337",bd=0,font=("times new roman",12)).place(x=150,y=0)
remove_buttton=Button(OptionFrame,text="Remove Event",command=add_link3,cursor="hand2",bg="blue",fg="#d77337",bd=0,font=("times new roman",12)).place(x=230,y=0)
lbl_user=Label(Frame1,text="Search the Zoom code of Subject:",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=0,y=140)
txt_user=Entry(Frame1,font=("times new roman",15),bg="lightgray")
txt_user.place(x=0,y=170,width=350,height=35)
lbl_user1=Label(Frame1,text="Enter the day",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=0,y=80)
txt_user1=Entry(Frame1,font=("times new roman",15),bg="lightgray")
txt_user1.place(x=0,y=105,width=350,height=35)
search_button=Button(Frame1,text="Search Link",command=searchfun,cursor="hand2",bg="white",fg="#d77337",bd=0,font=("times new roman",12)).place(x=0,y=210)

root.mainloop()
