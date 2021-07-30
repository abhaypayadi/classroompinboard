from openpyxl import *
from tkinter import *
wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/tuesday.xlsx')
sheet = wb.active
  
  
def excel():
      
    
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
  
  
   
  

  
  
def focus1(event):
    
    subjecte.focus_set()
  
  

def focus2(event):
    
    teachere.focus_set()
  
  

def focus3(event):
    
    stimee.focus_set()
  

def focus4(event):
 
    etimee.focus_set()
  
  

def focus5(event):
    
    linke.focus_set()
  
  

def focus6(event):
    
    periode.focus_set()

  
  

def insert():
      
   
    if (subjecte.get() == "" and
        teachere.get() == "" and
        stimee.get() == "" and
        etimee.get() == "" and
        linke.get() == "" and
        periode.get() == ""):
              
        print("empty input")
  
    else:
  
      
        current_row = int(periode.get())+int(1)
        
  
      
        sheet.cell(row=current_row , column=1).value = subjecte.get()
        sheet.cell(row=current_row , column=2).value = teachere.get()
        sheet.cell(row=current_row , column=3).value = stimee.get()
        sheet.cell(row=current_row , column=4).value = etimee.get()
        sheet.cell(row=current_row , column=5).value = linke.get()
        title=Label(Frame_name,text="Event add Successfully",font=("Impact",10,"bold"),fg="#d77337",bg="white").place(x=0,y=0)
        
       
  

        wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/tuesday.xlsx')
  
        
  
   
        
  
  

if __name__ == "__main__":
      

    root = Tk()
  

    root.configure(background='white')
  
    root.title("ClassRoom PinBoard")
  
 
    root.geometry("1199x600+100+50")
    root.resizable(False,False)
    Frame_name=Frame(root,bg="white")
    Frame_name.place(x=150,y=500,height=30,width=210)
   
  
    excel()
  

    
  

    subject = Label(root, text="Subject", bg="light green")
  

    teacher = Label(root, text="Teacher", bg="light green")
  

    stime = Label(root, text="Start Time", bg="light green")
  

    etime = Label(root, text="End Time", bg="light green")
  

    link = Label(root, text="Link", bg="light green")

    period = Label(root, text="period", bg="light green")

  
    
  
    subject.grid(row=1, column=0)
    teacher.grid(row=2, column=0)
    stime.grid(row=3, column=0)
    etime.grid(row=4, column=0)
    link.grid(row=5, column=0)
    period.grid(row=6, column=0)
    
  
    
    subjecte = Entry(root)
    teachere = Entry(root)
    stimee = Entry(root)
    etimee = Entry(root)
    linke = Entry(root)
    periode = Entry(root)
    
  
    
    subjecte.bind("<Return>", focus1)
  
   
    teachere.bind("<Return>", focus2)
  

    stimee.bind("<Return>", focus3)

    etimee.bind("<Return>", focus4)
  

    linke.bind("<Return>", focus5)
  
 
    periode.bind("<Return>", focus6)
  
  
    subjecte.grid(row=1, column=1, ipadx="100")
    teachere.grid(row=2, column=1, ipadx="100")
    stimee.grid(row=3, column=1, ipadx="100")
    etimee.grid(row=4, column=1, ipadx="100")
    linke.grid(row=5, column=1, ipadx="100")
    periode.grid(row=6, column=1, ipadx="100")

    

    excel()
  
 
    submit = Button(root, text="Submit", fg="Black",
                            bg="Red", command=insert)
    submit.grid(row=8, column=1)

  
 
    root.mainloop()