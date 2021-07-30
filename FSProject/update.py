from tkinter import *
from openpyxl import *
root=Tk()
root.title("ClassRoom Pinboard")
root.geometry("1199x600+100+50")
root.resizable(False,False)
Frame1=Frame(root,bg="white")
Frame1.place(x=150,y=0,height=600,width=1000)
def describe():
	extra=Label(Frame1,text="Select the day",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=450)
global extra	
lbl_user=Label(Frame1,text="Subject to be updated",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=140)
txt_user=Entry(Frame1,font=("times new roman",15),bg="lightgray")
txt_user.place(x=90,y=170,width=350,height=35)
lbl_user1=Label(Frame1,text="Updated Subject",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=225)
txt_user1=Entry(Frame1,font=("times new roman",15),bg="lightgray")
txt_user1.place(x=90,y=250,width=350,height=35)
lbl_user2=Label(Frame1,text="Updated Teacher",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=285)
txt_user2=Entry(Frame1,font=("times new roman",15),bg="lightgray")
txt_user2.place(x=90,y=310,width=350,height=35)
lbl_user3=Label(Frame1,text="Updated Start Time",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=350)
txt_user3=Entry(Frame1,font=("times new roman",15),bg="lightgray")
txt_user3.place(x=90,y=375,width=350,height=35)
lbl_user4=Label(Frame1,text="Updated End Time",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=610,y=225)
txt_user4=Entry(Frame1,font=("times new roman",15),bg="lightgray")
txt_user4.place(x=610,y=250,width=350,height=35)
lbl_user5=Label(Frame1,text="Updated Zoom Link",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=610,y=285)
txt_user5=Entry(Frame1,font=("times new roman",15),bg="lightgray")
txt_user5.place(x=610,y=310,width=350,height=35)
update_button=Button(Frame1,text="Update Subject",command=describe,cursor="hand2",bg="white",fg="#d77337",bd=0,font=("times new roman",12)).place(x=90,y=410)
my_menu=Menu(root)
root.config(menu=my_menu)
def mondayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/monday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)

			
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=450)
	else:
		sheet.cell(row=rowno,column=1).value=txt_user1.get()
		sheet.cell(row=rowno,column=2).value=txt_user2.get()
		sheet.cell(row=rowno,column=3).value=txt_user3.get()
		sheet.cell(row=rowno,column=4).value=txt_user4.get()
		sheet.cell(row=rowno,column=5).value=txt_user5.get()

	

	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/monday.xlsx')
def tuesdayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/tuesday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
		
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=450)
	else:
		sheet.cell(row=rowno,column=1).value=txt_user1.get()
		sheet.cell(row=rowno,column=2).value=txt_user2.get()
		sheet.cell(row=rowno,column=3).value=txt_user3.get()
		sheet.cell(row=rowno,column=4).value=txt_user4.get()
		sheet.cell(row=rowno,column=5).value=txt_user5.get()
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/tuesday.xlsx')	
def wednesdayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/wednesday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=450)
	else:
		sheet.cell(row=rowno,column=1).value=txt_user1.get()
		sheet.cell(row=rowno,column=2).value=txt_user2.get()
		sheet.cell(row=rowno,column=3).value=txt_user3.get()
		sheet.cell(row=rowno,column=4).value=txt_user4.get()
		sheet.cell(row=rowno,column=5).value=txt_user5.get()
		
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/wednesday.xlsx')	
def thursdayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/thursday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=450)
	else:
		sheet.cell(row=rowno,column=1).value=txt_user1.get()
		sheet.cell(row=rowno,column=2).value=txt_user2.get()
		sheet.cell(row=rowno,column=3).value=txt_user3.get()
		sheet.cell(row=rowno,column=4).value=txt_user4.get()
		sheet.cell(row=rowno,column=5).value=txt_user5.get()
		
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/thursday.xlsx')	
def fridayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/friday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=450)
	else:
		sheet.cell(row=rowno,column=1).value=txt_user1.get()
		sheet.cell(row=rowno,column=2).value=txt_user2.get()
		sheet.cell(row=rowno,column=3).value=txt_user3.get()
		sheet.cell(row=rowno,column=4).value=txt_user4.get()
		sheet.cell(row=rowno,column=5).value=txt_user5.get()
	

	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/friday.xlsx')
def saturdayfun():
	wb = load_workbook(r'/Users/abhaypayadi/Desktop/FSProject/dataset/saturday.xlsx')
	flag=0
	sheet = wb.active
	column_a=sheet['A']
	name=txt_user.get()
	for cell in column_a:
		if(cell.value==name):
			rowno=cell.row
			print(rowno)
			
			flag=1
	if(flag==0):
		extra=Label(Frame1,text="Invalid Subject name",font=("Goudy old style",15,"bold"),fg="gray",bg="white").place(x=90,y=450)
	else:
		sheet.cell(row=rowno,column=1).value=txt_user1.get()
		sheet.cell(row=rowno,column=2).value=txt_user2.get()
		sheet.cell(row=rowno,column=3).value=txt_user3.get()
		sheet.cell(row=rowno,column=4).value=txt_user4.get()
		sheet.cell(row=rowno,column=5).value=txt_user5.get()
		
	wb.save(r'/Users/abhaypayadi/Desktop/FSProject/dataset/saturday.xlsx')			
			
file_menu=Menu(my_menu)
my_menu.add_cascade(label="Days",menu=file_menu)
file_menu.add_command(label="Monday",command=mondayfun)
file_menu.add_command(label="Tuesday",command=tuesdayfun)
file_menu.add_command(label="Wednesday",command=wednesdayfun)
file_menu.add_command(label="Thursday",command=thursdayfun)
file_menu.add_command(label="Friday",command=fridayfun)
file_menu.add_command(label="Saturday",command=saturdayfun)


root.mainloop()